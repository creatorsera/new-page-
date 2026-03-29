"""
utils.py — MailHunter shared logic
All email helpers, validation engine, and XLSX builders live here.
Pages import what they need; nothing is duplicated.
"""
import re, io, smtplib, random, time
import requests
import streamlit as st
from datetime import datetime
from bs4 import BeautifulSoup
from email_validator import validate_email as ev_validate, EmailNotValidError
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import dns.resolver as _dns_resolver
    DNS_AVAILABLE = True
except ImportError:
    DNS_AVAILABLE = False

# ── CONSTANTS ─────────────────────────────────────────────────────────────────
EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", re.IGNORECASE)
TIER1 = re.compile(r"^(editor|admin|press|advert|contact)[a-z0-9._%+\-]*@", re.IGNORECASE)
TIER2 = re.compile(r"^(info|sales|hello|office|team|support|help)[a-z0-9._%+\-]*@", re.IGNORECASE)

BLOCKED_TLDS = {
    'png','jpg','jpeg','webp','gif','svg','ico','bmp','tiff','avif','mp4','mp3',
    'wav','ogg','mov','avi','webm','pdf','zip','rar','tar','gz','7z','js','css',
    'php','asp','aspx','xml','json','ts','jsx','tsx','woff','woff2','ttf','eot',
    'otf','map','exe','dmg','pkg','deb','apk',
}
PLACEHOLDER_DOMAINS = {
    'example.com','example.org','example.net','test.com','domain.com',
    'yoursite.com','yourwebsite.com','website.com','email.com','placeholder.com',
}
PLACEHOLDER_LOCALS = {
    'you','user','name','email','test','example','someone','username',
    'yourname','youremail','enter','address','sample',
}
SUPPRESS_PREFIXES = [
    'noreply','no-reply','donotreply','do-not-reply','mailer-daemon','bounce',
    'bounces','unsubscribe','notifications','notification','newsletter',
    'newsletters','postmaster','webmaster','auto-reply','autoreply','daemon',
    'robot','alerts','alert','system',
]
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Version/17 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64; rv:109.0) Gecko/20100101 Firefox/119.0",
    # mobile — helps with Facebook
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 Mobile/15E148 Safari/604.1",
]
FREE_EMAIL_DOMAINS = {
    "gmail.com","yahoo.com","hotmail.com","outlook.com","aol.com",
    "icloud.com","protonmail.com","zoho.com","live.com","msn.com",
}

_DISPOSABLE_FALLBACK = {
    'mailinator.com','guerrillamail.com','tempmail.com','throwaway.email','yopmail.com',
    'sharklasers.com','spam4.me','trashmail.com','trashmail.me','maildrop.cc',
    '10minutemail.com','fakeinbox.com','discard.email','mailnesia.com',
    'tempr.email','trashmail.at','trashmail.io','wegwerfemail.de','meltmail.com',
}

@st.cache_data(ttl=86400, show_spinner=False)
def fetch_disposable_domains():
    try:
        r = requests.get(
            "https://raw.githubusercontent.com/disposable-email-domains/"
            "disposable-email-domains/main/disposable_email_blocklist.conf",
            timeout=8)
        if r.status_code == 200:
            return set(r.text.splitlines())
    except Exception:
        pass
    return _DISPOSABLE_FALLBACK

# ── EMAIL HELPERS ─────────────────────────────────────────────────────────────
def is_valid_email(email):
    e = email.strip()
    if not e or e.count('@') != 1: return False
    local, domain = e.split('@'); lo, do = local.lower(), domain.lower()
    if not local or not domain: return False
    if local.startswith('.') or local.endswith('.') or local.startswith('-'): return False
    if len(local) > 64 or len(domain) > 255: return False
    if '.' not in domain: return False
    tld = do.rsplit('.',1)[-1]
    if len(tld) < 2 or tld in BLOCKED_TLDS: return False
    if re.search(r'@\d+x[\-\d]','@'+do): return False
    if re.match(r'^\d+x', do): return False
    if do in PLACEHOLDER_DOMAINS: return False
    if lo in PLACEHOLDER_LOCALS: return False
    if any(lo == p or lo.startswith(p) for p in SUPPRESS_PREFIXES): return False
    if re.search(r'\d+x\d+', lo): return False
    return True

def tier_key(e):
    if TIER1.match(e): return "1"
    if TIER2.match(e): return "2"
    return "3"

def tier_short(e): return {"1":"Tier 1","2":"Tier 2","3":"Tier 3"}[tier_key(e)]
def sort_by_tier(emails): return sorted(emails, key=tier_key)

def pick_best(emails):
    pool = [e for e in emails if is_valid_email(e)]
    if not pool: return None
    for pat in [TIER1, TIER2]:
        h = [e for e in pool if pat.match(e)]
        if h: return h[0]
    return pool[0]

def confidence_score(email, val):
    if not val: return None
    s = 100; t = tier_key(email)
    if t == "2": s -= 10
    if t == "3": s -= 25
    if not val.get("spf"):       s -= 15
    if val.get("catch_all"):     s -= 20
    if val.get("free"):          s -= 8
    st_ = val.get("status","")
    if st_ == "Risky":           s -= 30
    if st_ == "Not Deliverable": s -= 65
    return max(0, s)

def conf_color(sc):
    if sc is None: return "#ccc"
    if sc >= 75:   return "#16a34a"
    if sc >= 45:   return "#d97706"
    return "#dc2626"

def val_icon(s):
    return {"Deliverable":"✅","Risky":"⚠️","Not Deliverable":"❌"}.get(s,"—")

# ── HTTP ──────────────────────────────────────────────────────────────────────
def make_headers(mobile=False):
    pool = USER_AGENTS
    ua = pool[-1] if mobile else random.choice(pool[:3])
    return {
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
        "Accept-Language": "en-US,en;q=0.5",
    }

def fetch_page(url, timeout=10, mobile=False):
    """Returns (html_or_None, status_code)."""
    try:
        r = requests.get(url, headers=make_headers(mobile), timeout=timeout, allow_redirects=True)
        if "text" in r.headers.get("Content-Type","") and r.ok:
            return r.text, r.status_code
        return None, r.status_code
    except Exception:
        return None, 0

# ── EXTRACTION ────────────────────────────────────────────────────────────────
def extract_emails(html):
    soup = BeautifulSoup(html, "html.parser"); raw = set()
    raw.update(EMAIL_REGEX.findall(soup.get_text(" ")))
    raw.update(EMAIL_REGEX.findall(html))
    for a in soup.find_all("a", href=True):
        if a["href"].lower().startswith("mailto:"):
            raw.add(a["href"][7:].split("?")[0].strip())
    return {e for e in raw if is_valid_email(e)}

def extract_social(html):
    TWITTER_SKIP  = {'share','intent','home','search','hashtag','i','status','twitter','x'}
    LINKEDIN_SKIP = {'share','shareArticle','in','company','pub','feed','login','authwall'}
    FACEBOOK_SKIP = {'sharer','share','dialog','login','home','watch','groups','events','marketplace'}
    soup = BeautifulSoup(html,"html.parser"); tw,li,fb = set(),set(),set()
    for a in soup.find_all("a", href=True):
        href = a["href"]; hl = href.lower()
        if "twitter.com/" in hl or "x.com/" in hl:
            m = re.search(r'(?:twitter\.com|x\.com)/([A-Za-z0-9_]{1,50})', href)
            if m and m.group(1).lower() not in TWITTER_SKIP: tw.add("@"+m.group(1))
        elif "linkedin.com/in/" in hl or "linkedin.com/company/" in hl:
            m = re.search(r'linkedin\.com/(in|company)/([^/?&#\s]{2,80})', href)
            if m and m.group(2).lower() not in LINKEDIN_SKIP:
                li.add(f"linkedin.com/{m.group(1)}/{m.group(2)}")
        elif "facebook.com/" in hl and "facebook.com/tr?" not in hl:
            m = re.search(r'facebook\.com/([A-Za-z0-9_.]{3,80})', href)
            if m and m.group(1).lower() not in FACEBOOK_SKIP:
                fb.add(f"facebook.com/{m.group(1)}")
    return tw, li, fb

# ── VALIDATION ENGINE ─────────────────────────────────────────────────────────
def _val_syntax(email):
    try: ev_validate(email); return True
    except EmailNotValidError: return False

def _val_mx(domain):
    try:
        recs = _dns_resolver.resolve(domain,"MX")
        return True, [str(r.exchange) for r in recs]
    except: return False, []

def _val_spf(domain):
    try:
        for rd in _dns_resolver.resolve(domain,"TXT"):
            if "v=spf1" in str(rd): return True
    except: pass
    return False

def _val_dmarc(domain):
    try:
        for rd in _dns_resolver.resolve(f"_dmarc.{domain}","TXT"):
            if "v=DMARC1" in str(rd): return True
    except: pass
    return False

def _val_mailbox(email, mx_records):
    try:
        mx = mx_records[0].rstrip(".")
        with smtplib.SMTP(mx, timeout=6) as s:
            s.helo("example.com"); s.mail("test@example.com")
            code, _ = s.rcpt(email)
            return code == 250
    except: return False

def _val_catch_all(domain, mx_records):
    try:
        mx = mx_records[0].rstrip(".")
        with smtplib.SMTP(mx, timeout=6) as s:
            s.helo("example.com"); s.mail("test@example.com")
            code, _ = s.rcpt(f"randomaddress9x7z@{domain}")
            return code == 250
    except: return False

def _deliverability(syntax, domain_ok, mailbox_ok, disposable, free, catch_all, mx_ok, spf_ok):
    if not syntax:    return "Not Deliverable","Invalid syntax"
    if not domain_ok: return "Not Deliverable","Domain doesn't exist"
    if disposable:    return "Not Deliverable","Disposable domain"
    if not mx_ok:     return "Not Deliverable","No MX records"
    if mailbox_ok:
        if free: return ("Risky","Catch-all + free") if catch_all else ("Deliverable","Free provider")
        if catch_all:  return "Risky","Catch-all enabled"
        if not spf_ok: return "Risky","Missing SPF"
        return "Deliverable","—"
    else:
        if catch_all:  return "Risky","Catch-all, mailbox unknown"
        if free:       return "Deliverable","Free provider (unverified)"
        if not spf_ok: return "Risky","No SPF — spam risk"
        return "Deliverable","MX/SPF OK, mailbox unconfirmed"

def validate_email_full(email):
    disp = fetch_disposable_domains()
    domain = email.split("@")[-1].lower()
    syntax = _val_syntax(email)
    mx_ok, mx_h = _val_mx(domain) if DNS_AVAILABLE else (False,[])
    spf   = _val_spf(domain)   if DNS_AVAILABLE else False
    dmarc = _val_dmarc(domain) if DNS_AVAILABLE else False
    disp_ = domain in disp
    free  = domain in FREE_EMAIL_DOMAINS
    mbox  = _val_mailbox(email, mx_h) if (mx_ok and DNS_AVAILABLE) else False
    ca    = _val_catch_all(domain, mx_h) if (mx_ok and DNS_AVAILABLE) else False
    status, reason = _deliverability(syntax, mx_ok, mbox, disp_, free, ca, mx_ok, spf)
    return {"status":status,"reason":reason,"syntax":syntax,"mx":mx_ok,
            "spf":spf,"dmarc":dmarc,"mailbox":mbox,"disposable":disp_,"free":free,"catch_all":ca}

def validate_with_fallback(all_emails, current_best, existing_val=None):
    """
    Validates current_best. If not Deliverable, tries others in tier order.
    Returns (chosen_email, validation_result, was_fallback, original_status).
    """
    if not current_best or not all_emails:
        return current_best, None, False, None
    val = existing_val or validate_email_full(current_best)
    original_status = val["status"]
    if val["status"] == "Deliverable":
        return current_best, val, False, original_status
    # try others in tier order
    for email in sort_by_tier(all_emails):
        if email == current_best: continue
        v = validate_email_full(email)
        if v["status"] == "Deliverable":
            return email, v, True, original_status
    # settle for best Risky if current is Not Deliverable
    if val["status"] == "Not Deliverable":
        for email in sort_by_tier(all_emails):
            if email == current_best: continue
            v = validate_email_full(email)
            if v["status"] == "Risky":
                return email, v, True, original_status
    return current_best, val, False, original_status

# ══════════════════════════════════════════════════════════════════════════════
#  XLSX STYLE PALETTE — shared across all builders
# ══════════════════════════════════════════════════════════════════════════════
def _make_fill(hex_): return PatternFill("solid", fgColor=hex_)
def _make_font(bold=False, color="111111", size=10, name="Calibri", italic=False):
    return Font(bold=bold, color=color, size=size, name=name, italic=italic)
def _make_border():
    t = Side(style="thin", color="E5E7EB")
    return Border(left=t, right=t, top=t, bottom=t)
def _center(): return Alignment(horizontal="center", vertical="center")
def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=False)

# row fills
RF_DELIV = _make_fill("F0FDF4")
RF_RISKY = _make_fill("FFFBEB")
RF_BAD   = _make_fill("FFF1F2")
RF_NONE  = _make_fill("F9FAFB")
# email cell accents
EF_DELIV  = _make_fill("DCFCE7")
EF_RISKY  = _make_fill("FEF3C7")
EF_BAD    = _make_fill("FECACA")
EF_FALLBK = _make_fill("E0F2FE")   # sky — fallback email used
# tier fills
TF_T1 = _make_fill("FEF9C3")
TF_T2 = _make_fill("EEF2FF")
TF_T3 = _make_fill("F1F5F9")
# confidence fills
CF_HIGH = _make_fill("D1FAE5")
CF_MID  = _make_fill("FEF3C7")
CF_LOW  = _make_fill("FEE2E2")
# status fills (solid, for status cell)
SF = {
    "Deliverable":    _make_fill("16A34A"),
    "Risky":          _make_fill("D97706"),
    "Not Deliverable":_make_fill("DC2626"),
}
HDR_FILL = _make_fill("111111")

def _row_fill(status):
    return {"Deliverable":RF_DELIV,"Risky":RF_RISKY,"Not Deliverable":RF_BAD}.get(status, RF_NONE)

def _email_fill(status, was_fallback):
    if was_fallback: return EF_FALLBK
    return {"Deliverable":EF_DELIV,"Risky":EF_RISKY,"Not Deliverable":EF_BAD}.get(status)

def _tier_fill(tier):
    if "Tier 1" in tier: return TF_T1
    if "Tier 2" in tier: return TF_T2
    return TF_T3

def _conf_fill(score):
    if score is None: return None
    if score >= 75: return CF_HIGH
    if score >= 45: return CF_MID
    return CF_LOW

def _hdr(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = HDR_FILL
    c.font = _make_font(bold=True, color="FFFFFF", size=10)
    c.alignment = _center()
    c.border = _make_border()
    if width: ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _cell(ws, row, col, value, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:  c.fill  = fill
    if font:  c.font  = font
    if align: c.alignment = align
    c.border = _make_border()
    return c

def _stat_sheet(wb, sheet_name, stats_rows, title, subtitle=""):
    """Reusable stats sheet builder with unicode bar charts."""
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 30
    # title
    t = ws.cell(row=1, column=1, value=title)
    t.font = _make_font(bold=True, size=15); t.fill = _make_fill("F9FAFB")
    ws.merge_cells("A1:C1"); ws.row_dimensions[1].height = 28
    t.alignment = _left()
    if subtitle:
        s = ws.cell(row=2, column=1, value=subtitle)
        s.font = _make_font(color="999999", size=9, italic=True)
        ws.merge_cells("A2:C2")
    ts = ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}")
    ts.font = _make_font(color="AAAAAA", size=9)
    ws.merge_cells("A3:C3")

    COLORS = {
        "sites":"0C4A6E","emails":"0C4A6E",
        "tier1":"78350F","tier2":"3730A3","tier3":"334155",
        "deliverable":"14532D","risky":"78350F","fail":"881337",
        "fallback":"0C4A6E","none":"374151","avg":"14532D",
        "total":"0C4A6E","default":"374151",
    }
    BG = {
        "sites":"F0F9FF","emails":"F0F9FF",
        "tier1":"FEF9C3","tier2":"EEF2FF","tier3":"F1F5F9",
        "deliverable":"F0FDF4","risky":"FFFBEB","fail":"FFF1F2",
        "fallback":"E0F2FE","none":"F9FAFB","avg":"F0FDF4",
        "total":"F0F9FF","default":"F9FAFB",
    }

    total = max(1, next((v for _,v,_ in stats_rows if "total" in _.lower() or "sites" in _.lower()), 1))

    for i, (label, value, key) in enumerate(stats_rows, 5):
        fg = COLORS.get(key, COLORS["default"])
        bg = BG.get(key, BG["default"])
        fill_ = _make_fill(bg)
        lc = _cell(ws, i, 1, label, fill_, _make_font(color=fg, size=10), _left())
        vc = _cell(ws, i, 2, value, fill_, _make_font(color=fg, size=11, bold=True), _center())
        ws.row_dimensions[i].height = 21
        if isinstance(value, (int,float)) and key not in ("avg",):
            pct = min(float(value)/total, 1.0)
            n = int(pct*22)
            bar = "█"*n + "░"*(22-n) + f"  {round(pct*100)}%"
            bc = _cell(ws, i, 3, bar, fill_, _make_font(size=9, name="Courier New", color=fg), _left())
        else:
            _cell(ws, i, 3, "", fill_)
    return ws

# ══════════════════════════════════════════════════════════════════════════════
#  XLSX BUILDER — SCRAPER
# ══════════════════════════════════════════════════════════════════════════════
def build_xlsx_scraper(results):
    """3-sheet scraper export: Results, All Emails, Stats."""
    wb = Workbook()
    BDR = _make_border()

    # ── Sheet 1: Results ──────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Results"; ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 26
    COLS = [("#",4),("Domain",22),("Email",30),("Tier",9),("Status",16),
            ("Score",8),("Reason",22),("SPF",6),("DMARC",7),("Catch-all",10),
            ("Fallback?",10),("Twitter",18),("LinkedIn",22),
            ("All Emails",40),("Pages",7),("Time(s)",9),("Source URL",30)]
    for ci,(name,w) in enumerate(COLS,1): _hdr(ws1,1,ci,name,w)

    for ri,(domain,r) in enumerate(results.items(),2):
        val    = r.get("Validation",{}) or {}
        status = val.get("status","")
        was_fb = r.get("WasFallback",False)
        vbest  = r.get("ValidatedBestEmail","") or r.get("Best Email","")
        tier   = r.get("Best Tier","")
        conf   = r.get("Confidence")
        rf     = _row_fill(status); ef = _email_fill(status,was_fb)
        ws1.row_dimensions[ri].height = 17

        def bc(col,val_,fill_=None,font_=None,align_=None):
            return _cell(ws1,ri,col,val_,fill_ or rf, font_ or _make_font(), align_ or _left())

        bc(1, ri-1, align_=_center())
        bc(2, domain, font_=_make_font(bold=True))
        _cell(ws1,ri,3,vbest, ef or rf, _make_font(name="Courier New",size=9), _left())
        _cell(ws1,ri,4,tier, _tier_fill(tier) if tier else rf, _make_font(), _center())
        # solid status cell
        sf_ = SF.get(status); sf_font = _make_font(bold=True,color="FFFFFF") if sf_ else _make_font()
        _cell(ws1,ri,5,status or "—",sf_ or rf,sf_font,_center())
        _cell(ws1,ri,6,conf if conf is not None else "—",_conf_fill(conf) or rf,
              _make_font(bold=True),_center())
        bc(7, val.get("reason","—"))
        _cell(ws1,ri,8,"✓" if val.get("spf") else "✗",rf,
              _make_font(color="16A34A" if val.get("spf") else "DC2626",size=11),_center())
        _cell(ws1,ri,9,"✓" if val.get("dmarc") else "✗",rf,
              _make_font(color="16A34A" if val.get("dmarc") else "DC2626",size=11),_center())
        _cell(ws1,ri,10,"⚠" if val.get("catch_all") else "—",rf,
              _make_font(color="D97706" if val.get("catch_all") else "AAAAAA",size=11),_center())
        _cell(ws1,ri,11,"↻ Yes" if was_fb else "—",rf,
              _make_font(color="0891B2" if was_fb else "AAAAAA",bold=was_fb),_center())
        bc(12,(r.get("Twitter",[])+[""])[0])
        bc(13,(r.get("LinkedIn",[])+[""])[0])
        bc(14,"; ".join(r.get("All Emails",[])), font_=_make_font(size=8,name="Courier New",color="666666"))
        bc(15,r.get("Pages Scraped",0),align_=_center())
        bc(16,r.get("Total Time",""),align_=_center())
        bc(17,r.get("Source URL",""),font_=_make_font(color="2563EB",size=9))

    # legend
    leg_row = len(results)+3
    ws1.cell(row=leg_row,column=1,value="Legend:").font = _make_font(bold=True,size=9)
    for col,fill_,label in [(2,EF_DELIV,"Deliverable"),(4,EF_RISKY,"Risky"),
                             (6,EF_BAD,"Not Deliverable"),(8,EF_FALLBK,"Fallback used")]:
        c = ws1.cell(row=leg_row,column=col,value=label)
        c.fill=fill_; c.font=_make_font(size=9); c.alignment=_center()

    # ── Sheet 2: All Emails ───────────────────────────────────────────────
    ws2 = wb.create_sheet("All Emails"); ws2.freeze_panes="A2"
    for ci,(n,w) in enumerate([("Domain",20),("Email",30),("Tier",9),("Best?",8),("Validated Best?",16)],1):
        _hdr(ws2,1,ci,n,w)
    r2=2
    for domain,r in results.items():
        best_e=r.get("Best Email",""); vbest_e=r.get("ValidatedBestEmail","") or best_e
        for email in r.get("All Emails",[]):
            t=tier_key(email); tf={"1":TF_T1,"2":TF_T2,"3":TF_T3}.get(t,RF_NONE)
            is_vb = email==vbest_e; rf2=EF_DELIV if is_vb else RF_NONE
            _cell(ws2,r2,1,domain,rf2,_make_font(bold=is_vb),_left())
            _cell(ws2,r2,2,email,rf2,_make_font(name="Courier New",size=9),_left())
            _cell(ws2,r2,3,tier_short(email),tf,_make_font(),_center())
            _cell(ws2,r2,4,"★" if email==best_e else "",rf2,_make_font(color="D97706",size=12),_center())
            _cell(ws2,r2,5,"✓" if is_vb else "",rf2,_make_font(color="16A34A",bold=True,size=12),_center())
            ws2.row_dimensions[r2].height=15; r2+=1

    # ── Sheet 3: Stats ────────────────────────────────────────────────────
    n_sites=len(results); n_emails=sum(len(r.get("All Emails",[])) for r in results.values())
    n_t1=sum(1 for r in results.values() if r.get("Best Tier","").startswith("Tier 1"))
    n_del=sum(1 for r in results.values() if (r.get("Validation",{}) or {}).get("status")=="Deliverable")
    n_risk=sum(1 for r in results.values() if (r.get("Validation",{}) or {}).get("status")=="Risky")
    n_bad=sum(1 for r in results.values() if (r.get("Validation",{}) or {}).get("status")=="Not Deliverable")
    n_fb=sum(1 for r in results.values() if r.get("WasFallback"))
    n_none=sum(1 for r in results.values() if not r.get("Best Email"))
    vals_with_conf=[r.get("Confidence") for r in results.values() if r.get("Confidence") is not None]
    avg_conf=round(sum(vals_with_conf)/len(vals_with_conf),1) if vals_with_conf else "—"

    _stat_sheet(wb,"Stats",[
        ("Total sites scanned",  n_sites,  "sites"),
        ("Total emails found",   n_emails, "emails"),
        ("Tier 1 emails",        n_t1,     "tier1"),
        ("Validated Deliverable",n_del,    "deliverable"),
        ("Validated Risky",      n_risk,   "risky"),
        ("Not Deliverable",      n_bad,    "fail"),
        ("Fallback emails used", n_fb,     "fallback"),
        ("Sites with no email",  n_none,   "none"),
        ("Avg confidence score", avg_conf, "avg"),
    ], "MailHunter — Scraper Results", f"{n_sites} sites scanned")

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
#  XLSX BUILDER — FACEBOOK
# ══════════════════════════════════════════════════════════════════════════════
def build_xlsx_facebook(fb_results):
    """2-sheet Facebook extractor export: Results, All Emails."""
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Facebook Results"; ws1.freeze_panes="A2"
    ws1.row_dimensions[1].height = 26

    COLS = [("#",4),("Handle",22),("Page URL",34),("Emails Found",8),
            ("Best Email",30),("Tier",9),("Status",16),("All Emails",44),("Time(s)",9)]
    for ci,(n,w) in enumerate(COLS,1): _hdr(ws1,1,ci,n,w)

    for ri,(handle,r) in enumerate(fb_results.items(),2):
        emails = r.get("emails",[])
        best   = pick_best(emails) or ""
        tier   = tier_short(best) if best else ""
        status = r.get("status","scraped")
        t_fill = _tier_fill(tier) if tier else RF_NONE
        rf     = RF_DELIV if emails else RF_NONE
        ws1.row_dimensions[ri].height = 17

        _cell(ws1,ri,1,ri-1,rf,_make_font(),_center())
        _cell(ws1,ri,2,handle,rf,_make_font(bold=True),_left())
        _cell(ws1,ri,3,f"https://facebook.com/{handle}",rf,_make_font(color="2563EB",size=9),_left())
        _cell(ws1,ri,4,len(emails),rf,_make_font(bold=True),_center())
        _cell(ws1,ri,5,best,EF_DELIV if best else rf,_make_font(name="Courier New",size=9),_left())
        _cell(ws1,ri,6,tier,t_fill,_make_font(),_center())
        status_fill = _make_fill("F0FDF4") if emails else (_make_fill("FFF1F2") if status=="blocked" else RF_NONE)
        _cell(ws1,ri,7,status,status_fill,_make_font(bold=bool(emails)),_center())
        _cell(ws1,ri,8,"; ".join(emails),rf,_make_font(size=8,name="Courier New",color="666666"),_left())
        _cell(ws1,ri,9,r.get("time",""),rf,_make_font(),_center())

    # Sheet 2: all emails expanded
    ws2 = wb.create_sheet("All Emails"); ws2.freeze_panes="A2"
    for ci,(n,w) in enumerate([("Handle",20),("Email",30),("Tier",9),("Is Best?",8)],1):
        _hdr(ws2,1,ci,n,w)
    r2=2
    for handle,r in fb_results.items():
        emails=r.get("emails",[]); best=pick_best(emails) or ""
        for email in sort_by_tier(emails):
            t=tier_key(email); tf={"1":TF_T1,"2":TF_T2,"3":TF_T3}.get(t,RF_NONE)
            is_b=email==best
            _cell(ws2,r2,1,handle,EF_DELIV if is_b else RF_NONE,_make_font(bold=is_b),_left())
            _cell(ws2,r2,2,email,EF_DELIV if is_b else RF_NONE,_make_font(name="Courier New",size=9),_left())
            _cell(ws2,r2,3,tier_short(email),tf,_make_font(),_center())
            _cell(ws2,r2,4,"★" if is_b else "",RF_NONE,_make_font(color="D97706",size=12),_center())
            ws2.row_dimensions[r2].height=15; r2+=1

    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
#  XLSX BUILDER — VALIDATOR
# ══════════════════════════════════════════════════════════════════════════════
def build_xlsx_validator(val_results):
    """2-sheet validator export: Validated Results, Stats."""
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Validated Results"; ws1.freeze_panes="A2"
    ws1.row_dimensions[1].height = 26

    COLS = [("#",4),("Email",32),("Domain",22),("Source",12),("Status",16),
            ("Score",8),("Tier",9),("Reason",22),("SPF",6),("DMARC",7),
            ("Catch-all",10),("Fallback?",10),("Original Email",28)]
    for ci,(n,w) in enumerate(COLS,1): _hdr(ws1,1,ci,n,w)

    for ri,row in enumerate(val_results,2):
        val    = row.get("val",{}) or {}
        status = val.get("status","")
        was_fb = row.get("was_fallback",False)
        email  = row.get("email","")
        orig   = row.get("original_email","")
        conf   = row.get("confidence")
        rf=_row_fill(status); ef=_email_fill(status,was_fb)
        ws1.row_dimensions[ri].height=17

        _cell(ws1,ri,1,ri-1,rf,_make_font(),_center())
        _cell(ws1,ri,2,email,ef or rf,_make_font(name="Courier New",size=9,bold=True),_left())
        _cell(ws1,ri,3,row.get("domain",""),rf,_make_font(),_left())
        src_colors={"Scraper":"0891B2","Facebook":"1D4ED8","Manual":"374151","FB":"1D4ED8"}
        src=row.get("source",""); src_c=src_colors.get(src,"374151")
        _cell(ws1,ri,4,src,rf,_make_font(color=src_c,bold=True),_center())
        sf_=SF.get(status)
        _cell(ws1,ri,5,status or "—",sf_ or rf,_make_font(bold=bool(sf_),color="FFFFFF" if sf_ else "111111"),_center())
        _cell(ws1,ri,6,conf if conf is not None else "—",_conf_fill(conf) or rf,_make_font(bold=True),_center())
        _cell(ws1,ri,7,tier_short(email) if email else "—",_tier_fill(tier_short(email)) if email else rf,_make_font(),_center())
        _cell(ws1,ri,8,val.get("reason","—"),rf,_make_font(),_left())
        _cell(ws1,ri,9,"✓" if val.get("spf") else "✗",rf,
              _make_font(color="16A34A" if val.get("spf") else "DC2626",size=11),_center())
        _cell(ws1,ri,10,"✓" if val.get("dmarc") else "✗",rf,
              _make_font(color="16A34A" if val.get("dmarc") else "DC2626",size=11),_center())
        _cell(ws1,ri,11,"⚠" if val.get("catch_all") else "—",rf,
              _make_font(color="D97706" if val.get("catch_all") else "AAAAAA",size=11),_center())
        _cell(ws1,ri,12,"↻ Yes" if was_fb else "—",rf,
              _make_font(color="0891B2" if was_fb else "AAAAAA",bold=was_fb),_center())
        _cell(ws1,ri,13,orig if was_fb else "—",rf,_make_font(color="AAAAAA",size=9),_left())

    # stats
    n_total=len(val_results)
    n_del=sum(1 for r in val_results if (r.get("val",{}) or {}).get("status")=="Deliverable")
    n_risk=sum(1 for r in val_results if (r.get("val",{}) or {}).get("status")=="Risky")
    n_bad=sum(1 for r in val_results if (r.get("val",{}) or {}).get("status")=="Not Deliverable")
    n_fb=sum(1 for r in val_results if r.get("was_fallback"))
    confs=[r.get("confidence") for r in val_results if r.get("confidence") is not None]
    avg_c=round(sum(confs)/len(confs),1) if confs else "—"
    srcs={}
    for r in val_results: srcs[r.get("source","?")] = srcs.get(r.get("source","?"),0)+1

    src_rows=[(f"Source: {k}",v,"total") for k,v in sorted(srcs.items(),key=lambda x:-x[1])]
    _stat_sheet(wb,"Stats",[
        ("Total emails validated",n_total,"emails"),
        ("Deliverable",n_del,"deliverable"),
        ("Risky",n_risk,"risky"),
        ("Not Deliverable",n_bad,"fail"),
        ("Fallback emails used",n_fb,"fallback"),
        ("Avg confidence score",avg_c,"avg"),
    ]+src_rows,"MailHunter — Validator Results",f"{n_total} emails validated")

    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
#  XLSX BUILDER — COLLECTIVE (the big one)
# ══════════════════════════════════════════════════════════════════════════════
def build_xlsx_collective(scraper_results, fb_results, val_results):
    """
    6-sheet collective export combining all sources:
    1. Master       — best email per domain from any source, deduped
    2. Scraper      — full scraper data
    3. Facebook     — FB extractor data
    4. Validated    — all validated rows
    5. All Emails   — every unique email, source-tagged
    6. Stats        — combined dashboard
    """
    wb = Workbook()

    # ── Sheet 1: Master ───────────────────────────────────────────────────
    ws_m = wb.active; ws_m.title = "Master"; ws_m.freeze_panes="A2"
    ws_m.row_dimensions[1].height = 26

    COLS_M = [("#",4),("Domain",22),("Best Email",30),("Source",12),("Tier",9),
              ("Status",16),("Score",8),("Reason",22),("SPF",6),("DMARC",7),
              ("Fallback?",10),("All Emails (combined)",50)]
    for ci,(n,w) in enumerate(COLS_M,1): _hdr(ws_m,1,ci,n,w)

    # Merge by domain — scraper wins, FB supplements
    all_domains = {}
    for domain,r in (scraper_results or {}).items():
        all_domains[domain] = {"emails":set(r.get("All Emails",[])),
                                "source":"Scraper","scraper_row":r,"fb_row":None}
    for handle,r in (fb_results or {}).items():
        domain = handle  # handles are domain-ish
        fb_emails = set(r.get("emails",[]))
        if domain in all_domains:
            all_domains[domain]["emails"].update(fb_emails)
            all_domains[domain]["source"] = "Both"
            all_domains[domain]["fb_row"] = r
        else:
            all_domains[domain] = {"emails":fb_emails,"source":"Facebook",
                                    "scraper_row":None,"fb_row":r}

    # pull validation data from val_results
    val_by_email = {row.get("email",""): row for row in (val_results or [])}
    # also pull from scraper validation
    for domain,r in (scraper_results or {}).items():
        vb = r.get("ValidatedBestEmail","") or r.get("Best Email","")
        if vb and vb not in val_by_email:
            val_ = r.get("Validation",{}) or {}
            if val_:
                val_by_email[vb] = {
                    "email":vb,"domain":domain,"val":val_,
                    "confidence":r.get("Confidence"),
                    "was_fallback":r.get("WasFallback",False),
                    "original_email":r.get("Best Email",""),
                    "source":"Scraper",
                }

    # sort: Deliverable → Risky → Not Deliverable → unvalidated
    def sort_key(item):
        _,d = item
        vb = (d.get("scraper_row") or {}).get("ValidatedBestEmail","") or pick_best(d["emails"]) or ""
        vr = val_by_email.get(vb,{}).get("val",{}) or {}
        order = {"Deliverable":0,"Risky":1,"Not Deliverable":2}.get(vr.get("status",""),3)
        return (order, next(iter(d["emails"]),"") if d["emails"] else "zzz")

    sorted_domains = sorted(all_domains.items(), key=sort_key)

    for ri,(domain,d) in enumerate(sorted_domains,2):
        emails_list = sort_by_tier(d["emails"])
        best = pick_best(d["emails"]) or ""
        sr = d.get("scraper_row") or {}
        vb = sr.get("ValidatedBestEmail","") or best
        vr_row = val_by_email.get(vb,{})
        val_ = vr_row.get("val",{}) or {}
        status = val_.get("status","")
        was_fb = vr_row.get("was_fallback",False) or sr.get("WasFallback",False)
        conf   = vr_row.get("confidence") or sr.get("Confidence")
        tier   = tier_short(vb) if vb else ""
        source = d["source"]
        rf = _row_fill(status); ef = _email_fill(status,was_fb)
        ws_m.row_dimensions[ri].height=17

        src_c={"Scraper":"0891B2","Facebook":"1D4ED8","Both":"7C3AED"}.get(source,"374151")
        _cell(ws_m,ri,1,ri-1,rf,_make_font(),_center())
        _cell(ws_m,ri,2,domain,rf,_make_font(bold=True),_left())
        _cell(ws_m,ri,3,vb,ef or rf,_make_font(name="Courier New",size=9,bold=bool(vb)),_left())
        _cell(ws_m,ri,4,source,rf,_make_font(color=src_c,bold=True),_center())
        _cell(ws_m,ri,5,tier,_tier_fill(tier) if tier else rf,_make_font(),_center())
        sf_=SF.get(status)
        _cell(ws_m,ri,6,status or "—",sf_ or rf,_make_font(bold=bool(sf_),color="FFFFFF" if sf_ else "111111"),_center())
        _cell(ws_m,ri,7,conf if conf is not None else "—",_conf_fill(conf) or rf,_make_font(bold=True),_center())
        _cell(ws_m,ri,8,val_.get("reason","—"),rf,_make_font(),_left())
        _cell(ws_m,ri,9,"✓" if val_.get("spf") else ("✗" if val_ else "—"),rf,
              _make_font(color="16A34A" if val_.get("spf") else "DC2626",size=11) if val_ else _make_font(color="AAAAAA"),_center())
        _cell(ws_m,ri,10,"✓" if val_.get("dmarc") else ("✗" if val_ else "—"),rf,
              _make_font(color="16A34A" if val_.get("dmarc") else "DC2626",size=11) if val_ else _make_font(color="AAAAAA"),_center())
        _cell(ws_m,ri,11,"↻" if was_fb else "—",rf,
              _make_font(color="0891B2" if was_fb else "AAAAAA",bold=was_fb),_center())
        _cell(ws_m,ri,12,"; ".join(emails_list[:8]),rf,_make_font(size=8,name="Courier New",color="666666"),_left())

    # ── Sheet 2: Scraper Results ──────────────────────────────────────────
    if scraper_results:
        _fill_scraper_sheet(wb, scraper_results)

    # ── Sheet 3: Facebook Results ─────────────────────────────────────────
    if fb_results:
        _fill_fb_sheet(wb, fb_results)

    # ── Sheet 4: Validated ────────────────────────────────────────────────
    if val_results:
        _fill_val_sheet(wb, val_results)

    # ── Sheet 5: All Unique Emails ────────────────────────────────────────
    ws5 = wb.create_sheet("All Unique Emails"); ws5.freeze_panes="A2"
    for ci,(n,w) in enumerate([("Email",32),("Domain",22),("Tier",9),
                                 ("Source(s)",14),("Status",16),("Score",8)],1):
        _hdr(ws5,1,ci,n,w)
    seen_emails = {}  # email -> {domain, sources, status, conf}
    for domain,r in (scraper_results or {}).items():
        for email in r.get("All Emails",[]):
            if email not in seen_emails:
                seen_emails[email]={"domain":domain,"sources":set(),"status":"","conf":None}
            seen_emails[email]["sources"].add("Scraper")
    for handle,r in (fb_results or {}).items():
        for email in r.get("emails",[]):
            if email not in seen_emails:
                seen_emails[email]={"domain":handle,"sources":set(),"status":"","conf":None}
            seen_emails[email]["sources"].add("Facebook")
    for row in (val_results or []):
        email=row.get("email","")
        if email in seen_emails:
            seen_emails[email]["status"]=(row.get("val",{}) or {}).get("status","")
            seen_emails[email]["conf"]=row.get("confidence")
    # pull from scraper validation too
    for domain,r in (scraper_results or {}).items():
        vb=r.get("ValidatedBestEmail","") or r.get("Best Email","")
        if vb in seen_emails and not seen_emails[vb]["status"]:
            seen_emails[vb]["status"]=(r.get("Validation",{}) or {}).get("status","")
            seen_emails[vb]["conf"]=r.get("Confidence")

    email_rows=sorted(seen_emails.items(),
                      key=lambda x:({"Deliverable":0,"Risky":1,"Not Deliverable":2}.get(x[1]["status"],3),
                                    tier_key(x[0])))
    for ri2,(email,info) in enumerate(email_rows,2):
        status2=info["status"]; conf2=info["conf"]
        rf2=_row_fill(status2); ef2=_email_fill(status2,False)
        src_str="+".join(sorted(info["sources"]))
        src_c2={"Scraper":"0891B2","Facebook":"1D4ED8","Scraper+Facebook":"7C3AED"}.get(src_str,"374151")
        _cell(ws5,ri2,1,email,ef2 or rf2,_make_font(name="Courier New",size=9),_left())
        _cell(ws5,ri2,2,info["domain"],rf2,_make_font(),_left())
        _cell(ws5,ri2,3,tier_short(email),_tier_fill(tier_short(email)),_make_font(),_center())
        _cell(ws5,ri2,4,src_str,rf2,_make_font(color=src_c2,bold=True),_center())
        sf2=SF.get(status2)
        _cell(ws5,ri2,5,status2 or "—",sf2 or rf2,
              _make_font(bold=bool(sf2),color="FFFFFF" if sf2 else "111111"),_center())
        _cell(ws5,ri2,6,conf2 if conf2 is not None else "—",_conf_fill(conf2) or rf2,_make_font(bold=True),_center())
        ws5.row_dimensions[ri2].height=15

    # ── Sheet 6: Combined Stats ───────────────────────────────────────────
    n_scr=len(scraper_results or {})
    n_fb=len(fb_results or {})
    n_val=len(val_results or [])
    n_uniq=len(seen_emails)
    n_del=sum(1 for v in seen_emails.values() if v["status"]=="Deliverable")
    n_risk=sum(1 for v in seen_emails.values() if v["status"]=="Risky")
    n_bad=sum(1 for v in seen_emails.values() if v["status"]=="Not Deliverable")
    n_fallbk=sum(1 for r in (val_results or []) if r.get("was_fallback")) + \
             sum(1 for r in (scraper_results or {}).values() if r.get("WasFallback"))
    confs2=[v["conf"] for v in seen_emails.values() if v["conf"] is not None]
    avg2=round(sum(confs2)/len(confs2),1) if confs2 else "—"

    _stat_sheet(wb,"Stats",[
        ("Sites scraped",       n_scr,  "sites"),
        ("FB pages scraped",    n_fb,   "emails"),
        ("Emails validated",    n_val,  "total"),
        ("Total unique emails", n_uniq, "emails"),
        ("Deliverable",         n_del,  "deliverable"),
        ("Risky",               n_risk, "risky"),
        ("Not Deliverable",     n_bad,  "fail"),
        ("Fallback emails used",n_fallbk,"fallback"),
        ("Avg confidence score",avg2,   "avg"),
    ],"MailHunter — Combined Export",
      f"{n_scr} scraped  |  {n_fb} FB pages  |  {n_uniq} unique emails")

    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

# helpers used by collective builder
def _fill_scraper_sheet(wb, results):
    ws = wb.create_sheet("Scraper Results"); ws.freeze_panes="A2"
    COLS=[("#",4),("Domain",22),("Email",30),("Tier",9),("Status",16),
          ("Score",8),("Reason",22),("SPF",6),("DMARC",7),("Pages",7),("Time(s)",9)]
    for ci,(n,w) in enumerate(COLS,1): _hdr(ws,1,ci,n,w)
    for ri,(domain,r) in enumerate(results.items(),2):
        val=r.get("Validation",{}) or {}; status=val.get("status","")
        vbest=r.get("ValidatedBestEmail","") or r.get("Best Email","")
        tier=r.get("Best Tier",""); conf=r.get("Confidence")
        rf=_row_fill(status); ef=_email_fill(status,r.get("WasFallback",False))
        ws.row_dimensions[ri].height=17
        _cell(ws,ri,1,ri-1,rf,_make_font(),_center())
        _cell(ws,ri,2,domain,rf,_make_font(bold=True),_left())
        _cell(ws,ri,3,vbest,ef or rf,_make_font(name="Courier New",size=9),_left())
        _cell(ws,ri,4,tier,_tier_fill(tier) if tier else rf,_make_font(),_center())
        sf_=SF.get(status)
        _cell(ws,ri,5,status or "—",sf_ or rf,_make_font(bold=bool(sf_),color="FFFFFF" if sf_ else "111111"),_center())
        _cell(ws,ri,6,conf if conf is not None else "—",_conf_fill(conf) or rf,_make_font(bold=True),_center())
        _cell(ws,ri,7,val.get("reason","—"),rf,_make_font(),_left())
        _cell(ws,ri,8,"✓" if val.get("spf") else "✗",rf,
              _make_font(color="16A34A" if val.get("spf") else "DC2626",size=11),_center())
        _cell(ws,ri,9,"✓" if val.get("dmarc") else "✗",rf,
              _make_font(color="16A34A" if val.get("dmarc") else "DC2626",size=11),_center())
        _cell(ws,ri,10,r.get("Pages Scraped",0),rf,_make_font(),_center())
        _cell(ws,ri,11,r.get("Total Time",""),rf,_make_font(),_center())

def _fill_fb_sheet(wb, fb_results):
    ws = wb.create_sheet("Facebook Results"); ws.freeze_panes="A2"
    COLS=[("#",4),("Handle",22),("Emails Found",10),("Best Email",30),("Tier",9),("Status",14),("Time(s)",9)]
    for ci,(n,w) in enumerate(COLS,1): _hdr(ws,1,ci,n,w)
    for ri,(handle,r) in enumerate(fb_results.items(),2):
        emails=r.get("emails",[]); best=pick_best(emails) or ""
        tier=tier_short(best) if best else ""; status=r.get("status","")
        rf=RF_DELIV if emails else RF_NONE
        ws.row_dimensions[ri].height=17
        _cell(ws,ri,1,ri-1,rf,_make_font(),_center())
        _cell(ws,ri,2,handle,rf,_make_font(bold=True),_left())
        _cell(ws,ri,3,len(emails),rf,_make_font(bold=True),_center())
        _cell(ws,ri,4,best,EF_DELIV if best else rf,_make_font(name="Courier New",size=9),_left())
        _cell(ws,ri,5,tier,_tier_fill(tier) if tier else rf,_make_font(),_center())
        sf_=_make_fill("F0FDF4") if emails else (_make_fill("FFF1F2") if status=="blocked" else RF_NONE)
        _cell(ws,ri,6,status,sf_,_make_font(bold=bool(emails)),_center())
        _cell(ws,ri,7,r.get("time",""),rf,_make_font(),_center())

def _fill_val_sheet(wb, val_results):
    ws = wb.create_sheet("Validated"); ws.freeze_panes="A2"
    COLS=[("#",4),("Email",30),("Domain",22),("Source",10),("Status",16),
          ("Score",8),("Tier",9),("Reason",22),("Fallback?",10)]
    for ci,(n,w) in enumerate(COLS,1): _hdr(ws,1,ci,n,w)
    for ri,row in enumerate(val_results,2):
        val_=row.get("val",{}) or {}; status=val_.get("status","")
        email=row.get("email",""); was_fb=row.get("was_fallback",False)
        conf=row.get("confidence"); source=row.get("source","")
        rf=_row_fill(status); ef=_email_fill(status,was_fb)
        src_c={"Scraper":"0891B2","Facebook":"1D4ED8","Manual":"374151"}.get(source,"374151")
        sf_=SF.get(status)
        ws.row_dimensions[ri].height=17
        _cell(ws,ri,1,ri-1,rf,_make_font(),_center())
        _cell(ws,ri,2,email,ef or rf,_make_font(name="Courier New",size=9,bold=True),_left())
        _cell(ws,ri,3,row.get("domain",""),rf,_make_font(),_left())
        _cell(ws,ri,4,source,rf,_make_font(color=src_c,bold=True),_center())
        _cell(ws,ri,5,status or "—",sf_ or rf,_make_font(bold=bool(sf_),color="FFFFFF" if sf_ else "111111"),_center())
        _cell(ws,ri,6,conf if conf is not None else "—",_conf_fill(conf) or rf,_make_font(bold=True),_center())
        _cell(ws,ri,7,tier_short(email) if email else "—",_tier_fill(tier_short(email)) if email else rf,_make_font(),_center())
        _cell(ws,ri,8,val_.get("reason","—"),rf,_make_font(),_left())
        _cell(ws,ri,9,"↻ Yes" if was_fb else "—",rf,_make_font(color="0891B2" if was_fb else "AAAAAA",bold=was_fb),_center())
